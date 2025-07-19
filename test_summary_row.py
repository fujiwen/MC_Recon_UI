import pandas as pd
import os
from openpyxl import load_workbook

# 测试脚本，用于检查合计行的供应商名称是否为空

def check_excel_file(file_path):
    print(f"检查文件: {file_path}")
    
    # 加载Excel文件
    wb = load_workbook(file_path)
    ws = wb.active
    
    # 获取最大行数
    max_row = ws.max_row
    
    # 获取列索引（假设供应商名称是最后一列）
    supplier_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=3, column=col).value == '供应商名称':
            supplier_col = col
            break
    
    if supplier_col is None:
        print("未找到供应商名称列")
        return
    
    # 检查合计行（最后一行）的供应商名称
    summary_row = max_row
    summary_cell_value = ws.cell(row=summary_row, column=1).value
    
    if summary_cell_value == '合计':
        supplier_name = ws.cell(row=summary_row, column=supplier_col).value
        print(f"合计行的供应商名称: '{supplier_name}'")
        if supplier_name == '' or supplier_name is None:
            print("修改成功: 合计行的供应商名称为空")
        else:
            print("修改失败: 合计行的供应商名称不为空")
    else:
        print(f"未找到合计行，最后一行的第一列值为: {summary_cell_value}")

# 获取第一个Excel文件进行测试（排除临时文件）
excel_dir = os.path.join('供应商对账明细', '202507')
if os.path.exists(excel_dir):
    # 排除以~$开头的临时文件
    excel_files = [f for f in os.listdir(excel_dir) if f.endswith('.xlsx') and not f.startswith('~$')]
    if excel_files:
        file_path = os.path.join(excel_dir, excel_files[0])
        print(f"找到文件: {file_path}")
        check_excel_file(file_path)
    else:
        print(f"在 {excel_dir} 目录中未找到Excel文件")
else:
    print(f"目录 {excel_dir} 不存在")