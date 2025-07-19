import sys
import os
import pandas as pd
import numpy as np
import re
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
from openpyxl.worksheet.properties import PageSetupProperties
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QLabel, QPushButton, QTextEdit, QProgressBar, QFrame,
                             QFileDialog, QMessageBox, QListWidget, QListWidgetItem)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer, QRect
from PyQt5.QtGui import QFont, QPalette, QColor, QIcon
from PyQt5.QtWidgets import QDesktopWidget

class DataProcessThread(QThread):
    progress_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str)
    
    def __init__(self, input_files):
        super().__init__()
        self.input_files = input_files

    def format_mixed_text(self, text):
        if pd.isna(text):
            return text
        text = str(text)
        chinese_pattern = re.compile('[\u4e00-\u9fff]')
        match = chinese_pattern.search(text)
        if match:
            english_part = text[:match.start()].strip()
            chinese_part = text[match.start():].strip()
            if english_part and chinese_part:
                return f'{english_part}\n{chinese_part}'
        return text

    def run(self):
        try:
            # 创建日志目录
            if not os.path.exists('logs'):
                os.makedirs('logs')
            
            # 配置日志
            log_filename = os.path.join('logs', f'process_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
            logging.basicConfig(
                level=logging.INFO,
                format='%(asctime)s - %(levelname)s - %(message)s',
                handlers=[
                    logging.FileHandler(log_filename, encoding='utf-8'),
                    logging.StreamHandler()
                ]
            )
            
            all_final_data = []
            
            for input_file in self.input_files:
                self.progress_signal.emit(f'开始读取文件：{os.path.basename(input_file)}')
                logging.info(f'开始读取文件：{input_file}')
                
                # 读取原始文件
                df = pd.read_excel(input_file, skiprows=8)
                logging.info(f'文件读取完成，共{len(df)}行数据')
                self.progress_signal.emit(f'文件读取完成，共{len(df)}行数据')
                
                # 获取收货单号的行索引
                receipt_rows = df[df['Unnamed: 0'].astype(str).str.match(r'^(RTS)?000\d+$', na=False)].index
                
                # 创建一个空的列表来存储所有明细数据
                all_details = []
                
                # 遍历每个收货单号之间的行
                total_receipts = len(receipt_rows)
                for i in range(total_receipts):
                    start_idx = receipt_rows[i]
                    end_idx = receipt_rows[i+1] if i < len(receipt_rows)-1 else len(df)
                    
                    receipt = df.loc[start_idx, 'Unnamed: 0']
                    supplier = df.loc[start_idx, 'Unnamed: 3']
                    date = df.loc[start_idx, 'Unnamed: 25']
                    
                    # 清理供应商名称和日期中的发票信息
                    if pd.notna(supplier):
                        supplier = re.sub(r'[（(].*[)）]|（专票.*|（普票.*|\s+专票.*|\s+普票.*|\d+%$', '', str(supplier)).strip()
                    
                    if pd.notna(date):
                        date = pd.to_datetime(date).strftime('%Y-%m-%d')
                    
                    # 获取明细行（跳过收货单号行）
                    details = df.loc[start_idx+1:end_idx-1].copy()
                    
                    # 只保留非空行且不包含Page和Delivery Date的行
                    details = details[details['Unnamed: 0'].notna()]
                    details = details[~details['Unnamed: 0'].astype(str).str.contains('Page|Delivery Date', na=False)]
                    
                    if not details.empty:
                        details['收货单号'] = receipt
                        details['供应商名称'] = supplier
                        details['收货日期'] = date
                        details['商品名称'] = details['Unnamed: 0'].apply(self.format_mixed_text)
                        details['实收数量'] = details['Unnamed: 9']
                        details['基本单位'] = details['Unnamed: 11']
                        details['单价'] = details['Unnamed: 15']
                        details['小计金额'] = details['Unnamed: 27']
                        details['税额'] = details['Unnamed: 32']
                        details['税率'] = details['Unnamed: 32'] / details['Unnamed: 27']
                        details['小计价税'] = details['Unnamed: 37']
                        details['部门'] = details['Unnamed: 39'].apply(self.format_mixed_text)
                        
                        all_details.append(details[['收货单号', '收货日期', '商品名称', '实收数量', '基本单位',
                                                   '单价', '小计金额', '税额', '税率', '小计价税', '部门', '供应商名称']])
                    
                    progress = f'处理进度：{i+1}/{total_receipts}'
                    self.progress_signal.emit(progress)
                    logging.info(progress)
                
                # 合并所有明细数据
                if all_details:
                    file_df = pd.concat(all_details, ignore_index=True)
                    all_final_data.append(file_df)
                    logging.info(f'文件处理完成，共整理{len(file_df)}条记录')
                    self.progress_signal.emit(f'文件处理完成，共整理{len(file_df)}条记录')
            
            # 合并所有文件的数据
            final_df = pd.concat(all_final_data, ignore_index=True)
            logging.info(f'所有文件处理完成，共整理{len(final_df)}条记录')
            self.progress_signal.emit(f'所有文件处理完成，共整理{len(final_df)}条记录')
            
            # 创建供应商对账明细表文件夹
            if not os.path.exists('供应商对账明细'):
                os.makedirs('供应商对账明细')
                logging.info('创建供应商对账明细文件夹')
            
            # 按供应商名称分组并生成对账明细表
            total_suppliers = len(final_df['供应商名称'].unique())
            current_supplier = 0
            
            for supplier_name, supplier_data in final_df.groupby('供应商名称'):
                if pd.notna(supplier_name) and supplier_name.strip():
                    current_supplier += 1
                    self.progress_signal.emit(f'正在生成供应商对账单 ({current_supplier}/{total_suppliers}): {supplier_name}')
                    
                    # 按收货日期和收货单号排序
                    supplier_data = supplier_data.sort_values(['收货日期', '收货单号'])
                    
                    # 获取年月信息
                    first_date = pd.to_datetime(supplier_data['收货日期'].iloc[0])
                    year_month = first_date.strftime('%Y%m')
                    
                    # 创建年月目录
                    year_month_dir = os.path.join('供应商对账明细', year_month)
                    if not os.path.exists(year_month_dir):
                        os.makedirs(year_month_dir)
                    
                    # 计算合计金额
                    total_amount = round(supplier_data['小计价税'].sum(), 2)
                    
                    # 创建一个包含合计行的新数据框
                    summary_row = pd.DataFrame([{
                        '收货单号': '合计',
                        '收货日期': '',
                        '商品名称': '',
                        '实收数量': '',
                        '基本单位': '',
                        '单价': '',
                        '小计金额': supplier_data['小计金额'].sum(),
                        '税额': supplier_data['税额'].sum(),
                        '税率': '',
                        '小计价税': total_amount,
                        '部门': '',
                        '供应商名称': ''
                    }])
                    
                    supplier_data_with_summary = pd.concat([supplier_data, summary_row], ignore_index=True)
                    
                    # 创建新的Excel工作簿
                    wb = Workbook()
                    ws = wb.active
                    
                    # 设置页面布局
                    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
                    ws.page_setup.paperSize = ws.PAPERSIZE_A4
                    ws.page_setup.fitToPage = True
                    ws.page_setup.fitToHeight = 0
                    ws.page_setup.fitToWidth = 1
                    ws.print_options.horizontalCentered = True
                    ws.print_options.verticalCentered = False
                    # 设置页脚文本、字体和大小
                    ws.oddFooter.center.text = '\n第 &P 页，共 &N 页\nSofitel Sanya Leeman Resort'
                    ws.oddFooter.center.size = 11
                    ws.oddFooter.center.font = '微软雅黑'

                    
                    # 设置页边距（单位：厘米）
                    ws.page_margins = PageMargins(left=0.31, right=0.31, top=0.31, bottom=0.39, header=0.31, footer=0.11)
                    
                    # 设置列宽
                    column_widths = {
                        '收货单号': 15,
                        '收货日期': 15,
                        '商品名称': 45,
                        '实收数量': 10,
                        '基本单位': 13,
                        '单价': 12,
                        '小计金额': 12,
                        '税额': 12,
                        '税率': 10,
                        '小计价税': 12,
                        '部门': 35,
                        '供应商名称': 36
                    }
                    
                    # 设置酒店名称标题
                    hotel_title_row = 1
                    ws.merge_cells(start_row=hotel_title_row, start_column=1, end_row=hotel_title_row, end_column=len(column_widths))
                    hotel_title_cell = ws.cell(row=hotel_title_row, column=1, value='对账明细表')
                    hotel_title_cell.font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
                    hotel_title_cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                    hotel_title_cell.alignment = Alignment(horizontal='center', vertical='center')
                    ws.row_dimensions[hotel_title_row].height = 60
                    
                    # 设置对账明细表标题
                    title_row = 2
                    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=len(column_widths))
                    title_cell = ws.cell(row=title_row, column=1, value='')
                    title_cell.font = Font(name='微软雅黑', size=20, bold=True, color='FFFFFF')
                    title_cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                    title_cell.alignment = Alignment(horizontal='center', vertical='center')
                    ws.row_dimensions[title_row].height = 10
                    
                    # 设置表头样式
                    header_font = Font(name='微软雅黑', size=13, bold=True, color='FFFFFF')
                    cell_font = Font(name='微软雅黑', size=13)
                    
                    # 设置对齐方式
                    center_alignment = Alignment(horizontal='center', vertical='center')
                    right_alignment = Alignment(horizontal='right', vertical='center', shrink_to_fit=False)
                    wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # 设置边框样式
                    thin_border = Border(
                        left=Side(style='hair', color='D3D3D3'),
                        right=Side(style='hair', color='D3D3D3'),
                        top=Side(style='hair', color='D3D3D3'),
                        bottom=Side(style='hair', color='D3D3D3')
                    )
                    thick_border = Border(
                        left=Side(style='thin', color='1F497D'),
                        right=Side(style='thin', color='1F497D'),
                        top=Side(style='thin', color='1F497D'),
                        bottom=Side(style='thin', color='1F497D')
                    )

                    # 写入表头
                    headers = list(supplier_data.columns)
                    header_row = 3
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=header_row, column=col, value=header)
                        cell.font = header_font
                        cell.alignment = center_alignment
                        cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                        cell.border = thick_border
                        ws.column_dimensions[get_column_letter(col)].width = column_widths[header]

                    # 写入数据
                    for row_idx, row in enumerate(supplier_data.values, header_row + 1):
                        # 设置行高为40
                        ws.row_dimensions[row_idx].height = 40
                        
                        # 检查是否为负数金额行
                        has_negative = False
                        for col_idx, value in enumerate(row, 1):
                            if headers[col_idx-1] in ['小计金额', '税额', '小计价税'] and pd.notna(value) and float(value) < 0:
                                has_negative = True
                                break
                        
                        # 设置斑马线效果（偶数行）
                        if row_idx % 2 == 0 and not has_negative:
                            row_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                        else:
                            row_fill = None
                        
                        # 写入单元格数据
                        for col_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                            cell.font = cell_font
                            cell.border = thin_border
                            
                            # 如果是负数金额行，整行设置黄色背景
                            if has_negative:
                                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                if headers[col_idx-1] in ['小计金额', '税额', '小计价税'] and pd.notna(value) and float(value) < 0:
                                    cell.font = Font(name='微软雅黑', size=11, color='FF0000')
                            elif row_fill:
                                cell.fill = row_fill
                            
                            # 设置数字列的对齐方式和格式
                            if headers[col_idx-1] in ['商品名称', '部门']:
                                cell.alignment = wrap_alignment
                            elif headers[col_idx-1] in ['实收数量', '单价', '小计金额', '税额', '小计价税']:
                                cell.alignment = right_alignment
                                if pd.notna(value) and str(value).strip():
                                    cell.number_format = '#,##0.00'
                            elif headers[col_idx-1] == '税率':
                                cell.alignment = right_alignment
                                if pd.notna(value) and str(value).strip():
                                    cell.number_format = '0%'
                            else:
                                cell.alignment = center_alignment
                    
                    # 写入合计行
                    row_idx = len(supplier_data) + header_row + 1
                    for col_idx, value in enumerate(summary_row.iloc[0], 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                        cell.border = thick_border
                        
                        # 设置数字列的对齐方式和格式
                        if headers[col_idx-1] in ['小计金额', '税额', '小计价税']:
                            cell.alignment = right_alignment
                            if pd.notna(value) and str(value).strip():
                                cell.number_format = '#,##0.00'
                        else:
                            cell.alignment = center_alignment
                    
                    # 设置重复打印的行

                    # 设置重复打印的行
                    ws.print_title_rows = '1:3'
                    
                    # 保存文件
                    output_file = os.path.join(year_month_dir, f'{supplier_name}_对账明细.xlsx')
                    wb.save(output_file)
                    logging.info(f'已生成供应商对账单：{output_file}')
            
            # 创建备份文件夹
            if not os.path.exists('bak'):
                os.makedirs('bak')
                logging.info('创建备份文件夹')
            
            # 获取当前时间作为备份文件名
            current_time = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
            
            # 备份数据
            backup_file = os.path.join('bak', f'cleaned_receiving_journal_{current_time}.xlsx')
            final_df.to_excel(backup_file, index=False)
            logging.info(f'数据已备份至：{backup_file}')
            
            self.progress_signal.emit('处理完成！')
            self.finished_signal.emit(True, '')
            
        except Exception as e:
            error_msg = f'处理过程中出现错误：{str(e)}'
            logging.error(error_msg)
            self.progress_signal.emit(error_msg)
            self.finished_signal.emit(False, error_msg)

class QTextEditLogger(logging.Handler):
    def __init__(self, widget):
        super().__init__()
        self.widget = widget
        self.widget.setReadOnly(True)
        self.widget.setFont(QFont('Helvetica', 16))  # 使用Helvetica字体
        
        # 设置样式
        self.widget.setStyleSheet("""
            QTextEdit {
                background-color: #2b2b2b;
                color: #ffffff;
                border: 1px solid #3c3c3c;
                border-radius: 5px;
                padding: 5px;
            }
        """)
        
        # 创建定时器用于更新日志
        self.update_timer = QTimer()
        self.update_timer.timeout.connect(self.update_log)
        self.update_timer.start(100)  # 每100ms更新一次
        self.pending_messages = []

    def emit(self, record):
        msg = self.format(record)
        self.pending_messages.append(msg)

    def update_log(self):
        if self.pending_messages:
            for msg in self.pending_messages:
                self.widget.append(msg)
            self.pending_messages.clear()
            # 滚动到底部
            self.widget.verticalScrollBar().setValue(
                self.widget.verticalScrollBar().maximum()
            )

# 程序版本信息
VERSION = '1.1.16'

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.selected_files = []
        self.version = VERSION
        self.initUI()
        
        # 记录应用程序启动日志
        logging.info(f"应用程序启动，版本：{self.version}")
        
    def centerOnScreen(self):
        """将窗口居中显示在屏幕上"""
        # 获取屏幕几何信息
        screen_geometry = QDesktopWidget().availableGeometry()
        # 计算窗口居中位置
        x = (screen_geometry.width() - self.width()) // 2
        y = (screen_geometry.height() - self.height()) // 2
        # 移动窗口到居中位置
        self.move(x, y)
    

    def initUI(self):
        self.setWindowTitle(f'MC对帐明细工具 v{self.version}')
        # 设置全局窗口图标
        icon = QIcon(':/icons/app_icon')
        QApplication.setWindowIcon(icon)
        
        # 设置窗口大小
        self.resize(1024, 768)
        
        # 窗口居中显示
        self.centerOnScreen()
        
        # 创建主窗口部件和布局
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        
        # 创建主布局
        layout = QVBoxLayout()
        
        # 创建左右分栏布局
        split_layout = QHBoxLayout()
        
        # 左侧：文件选择
        file_frame = QFrame()
        file_frame.setFrameShape(QFrame.StyledPanel)
        file_frame.setFrameShadow(QFrame.Raised)
        file_frame.setStyleSheet("""
            QFrame {
                background-color: #ffffff;
                border-radius: 10px;
                padding: 15px;
                margin: 10px;
            }
        """)
        
        file_layout = QVBoxLayout()
        
        # 文件选择标题和按钮区域
        header_layout = QHBoxLayout()
        self.file_label = QLabel('已选择的文件：')
        self.file_label.setProperty('title', 'true')
        self.select_button = QPushButton('添加文件')
        self.select_button.setStyleSheet("""
            QPushButton {
                background-color: #4a90e2;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #357abd;
            }
            QPushButton:pressed {
                background-color: #2a5f9e;
            }
        """)
        self.select_button.clicked.connect(self.selectFiles)
        
        # 添加清空选择按钮
        self.clear_button = QPushButton('清空选择')
        self.clear_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:pressed {
                background-color: #a93226;
            }
        """)
        self.clear_button.clicked.connect(self.clearFiles)
        
        header_layout.addWidget(self.file_label)
        header_layout.addStretch()
        header_layout.addWidget(self.clear_button)
        header_layout.addWidget(self.select_button)
        
        # 文件列表
        self.file_list = QListWidget()
        self.file_list.setStyleSheet("""
            QListWidget {
                border: 1px solid #e0e0e0;
                border-radius: 5px;
                padding: 5px;
                background-color: #f8f9fa;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #e0e0e0;
            }
            QListWidget::item:selected {
                background-color: #e3f2fd;
                color: #1976d2;
            }
        """)
        
        file_layout.addLayout(header_layout)
        file_layout.addWidget(self.file_list)
        file_frame.setLayout(file_layout)
        
        # 右侧：进度显示
        progress_frame = QFrame()
        progress_frame.setFrameShape(QFrame.StyledPanel)
        progress_frame.setFrameShadow(QFrame.Raised)
        progress_frame.setStyleSheet("""
            QFrame {
                background-color: #ffffff;
                border-radius: 10px;
                padding: 15px;
                margin: 10px;
            }
        """)
        
        progress_layout = QVBoxLayout()
        progress_label = QLabel('处理进度')
        progress_label.setProperty('title', 'true')
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: none;
                border-radius: 10px;
                text-align: center;
                background-color: #f0f0f0;
                height: 20px;
            }
            QProgressBar::chunk {
                background-color: #4caf50;
                border-radius: 10px;
            }
        """)
        self.progress_bar.setTextVisible(False)
        
        self.process_button = QPushButton('开始处理')
        self.process_button.setStyleSheet("""
            QPushButton {
                background-color: #4caf50;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                margin-top: 10px;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #43a047;
            }
            QPushButton:pressed {
                background-color: #388e3c;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.process_button.clicked.connect(self.startProcess)
        self.process_button.setEnabled(False)
        
        progress_layout.addWidget(progress_label)
        progress_layout.addWidget(self.progress_bar)
        progress_layout.addWidget(self.process_button)
        progress_layout.addStretch()
        progress_frame.setLayout(progress_layout)
        
        # 设置左右分栏的比例（5:5）
        split_layout.addWidget(file_frame, 5)
        split_layout.addWidget(progress_frame, 5)
        
        # 下方：日志显示
        log_frame = QFrame()
        log_frame.setFrameShape(QFrame.StyledPanel)
        log_frame.setFrameShadow(QFrame.Raised)
        log_frame.setStyleSheet("""
            QFrame {
                background-color: #ffffff;
                border-radius: 10px;
                padding: 15px;
                margin: 10px;
            }
        """)
        
        log_layout = QVBoxLayout()
        log_label = QLabel('处理日志')
        log_label.setProperty('title', 'true')
        self.progress_text = QTextEdit()
        self.progress_text.setReadOnly(True)
        
        log_layout.addWidget(log_label)
        log_layout.addWidget(self.progress_text)
        log_frame.setLayout(log_layout)
        
        # 添加所有部件到主布局（调整顺序，将日志放到下方）
        layout.addLayout(split_layout)
        layout.addWidget(log_frame)
        
        # 添加版权信息
        copyright_label = QLabel(f'Powered By Cayman Fu 2025 Ver {self.version}')
        copyright_label.setAlignment(Qt.AlignCenter)
        copyright_label.setStyleSheet("color: #666666;")
        layout.addWidget(copyright_label)
        
        main_widget.setLayout(layout)
        
        # 设置整体样式
        self.setStyleSheet("""
            * {
                font-family: "微软雅黑";
                font-size: 16px;
            }
            QPushButton {
                font-size: 14px;
            }
            QMainWindow {
                background-color: #f0f2f5;
            }
            QLabel {
                font-size: 16px;
            }
            QLabel[title="true"] {
                font-weight: bold;
                font-size: 16px;
            }
            QPushButton {
                font-size: 16px;
            }
            QListWidget, QListWidget::item {
                font-size: 16px;
            }
            QTextEdit {
                font-size: 16px;
            }
            QProgressBar {
                font-size: 16px;
            }
            QMessageBox {
                font-size: 16px;
            }
            QMessageBox QLabel {
                font-size: 16px;
            }
            QMessageBox QPushButton {
                font-size: 16px;
                min-width: 70px;
                padding: 6px 12px;
            }
            QFileDialog {
                font-size: 16px;
            }
            QFileDialog QLabel {
                font-size: 16px;
            }
            QFileDialog QPushButton {
                font-size: 16px;
            }
            QFileDialog QComboBox {
                font-size: 16px;
            }
            QFileDialog QListView {
                font-size: 16px;
            }
        """)
    
    def selectFiles(self):
        # 尝试从上次的位置打开文件对话框
        last_dir = getattr(self, 'last_directory', '')
        if not last_dir or not os.path.exists(last_dir):
            last_dir = ''
            
        files, _ = QFileDialog.getOpenFileNames(
            self,
            '选择文件',
            last_dir,
            'Excel Files (*.xls *.xlsx);;All Files (*)'
        )
        if files:
            # 记住最后打开的目录
            self.last_directory = os.path.dirname(files[0])
            logging.info(f'选择了{len(files)}个文件')
            
            # 避免重复添加相同的文件
            new_files = [f for f in files if f not in self.selected_files]
            if new_files:
                self.selected_files.extend(new_files)
                self.updateFileList()
                self.process_button.setEnabled(True)
            else:
                warning_box = QMessageBox(self)
                warning_box.setWindowTitle('警告')
                warning_box.setText('所选文件已存在！')
                warning_box.setIcon(QMessageBox.Warning)
                warning_box.exec_()
    
    def clearFiles(self):
        """清空文件列表并重置界面状态"""
        self.selected_files.clear()
        self.updateFileList()
        self.process_button.setEnabled(False)
        logging.info('已清空文件列表')
    
    def updateFileList(self):
        self.file_list.clear()
        for file_path in self.selected_files:
            self.file_list.addItem(QListWidgetItem(file_path))
    
    def startProcess(self):
        if not self.selected_files:
            warning_box = QMessageBox(self)
            warning_box.setWindowTitle('警告')
            warning_box.setText('请先选择要处理的文件！')
            warning_box.setIcon(QMessageBox.Warning)
            warning_box.exec_()
            return
        
        self.process_button.setEnabled(False)
        self.select_button.setEnabled(False)
        self.clear_button.setEnabled(False)
        self.progress_text.clear()
        self.progress_bar.setRange(0, 0)  # 设置进度条为忙碌状态
        
        # 创建并启动处理线程
        self.process_thread = DataProcessThread(self.selected_files)
        self.process_thread.progress_signal.connect(self.updateProgress)
        self.process_thread.finished_signal.connect(self.processFinished)
        self.process_thread.start()
    
    def updateProgress(self, message):
        self.progress_text.append(message)
        # 滚动到底部
        self.progress_text.verticalScrollBar().setValue(
            self.progress_text.verticalScrollBar().maximum()
        )
    
    def processFinished(self, success, error_msg):
        self.progress_bar.setRange(0, 100)  # 恢复进度条正常状态
        self.progress_bar.setValue(100 if success else 0)
        self.process_button.setEnabled(True)
        self.select_button.setEnabled(True)
        self.clear_button.setEnabled(True)
        
        if success:
            # 获取处理的统计信息
            supplier_dir = '供应商对账明细'
            year_month_dirs = [d for d in os.listdir(supplier_dir) if os.path.isdir(os.path.join(supplier_dir, d))]
            
            if year_month_dirs:
                latest_dir = max(year_month_dirs)  # 获取最新的年月目录
                full_dir_path = os.path.join(supplier_dir, latest_dir)
                supplier_files = [f for f in os.listdir(full_dir_path) if f.endswith('.xlsx') and not f.startswith('~$')]
                
                stats_message = f'数据处理完成！\n\n处理结果:\n- 生成了{len(supplier_files)}个供应商对账单\n- 保存在目录: {full_dir_path}\n\n是否打开输出文件夹？'
            else:
                stats_message = '数据处理完成！是否打开输出文件夹？'
            
            info_box = QMessageBox(self)
            info_box.setWindowTitle('完成')
            info_box.setText(stats_message)
            info_box.setIcon(QMessageBox.Information)
            info_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            info_box.setDefaultButton(QMessageBox.Yes)
            reply = info_box.exec_()
            
            if reply == QMessageBox.Yes:
                # 使用跨平台的方法打开文件夹
                folder_path = os.path.abspath('供应商对账明细')
                try:
                    import subprocess
                    import webbrowser
                    
                    # 首先尝试使用平台特定的方法
                    if sys.platform == 'win32':
                        os.startfile(folder_path)  # Windows特有方法
                    elif sys.platform == 'darwin':  # macOS
                        subprocess.Popen(['open', folder_path])
                    elif sys.platform.startswith('linux'):  # Linux
                        subprocess.Popen(['xdg-open', folder_path])
                    else:
                        # 如果以上都不适用，尝试使用webbrowser模块
                        webbrowser.open('file://' + folder_path)
                except Exception as e:
                    logging.warning(f'无法打开文件夹：{e}')
                    warning_box = QMessageBox(self)
                    warning_box.setWindowTitle('提示')
                    warning_box.setText(f'无法自动打开文件夹，请手动查看：{folder_path}')
                    warning_box.setIcon(QMessageBox.Warning)
                    warning_box.exec_()
            # 处理完成后自动清空文件列表
            self.clearFiles()
            logging.info('处理完成，界面已重置')
        else:
            error_box = QMessageBox(self)
            error_box.setWindowTitle('错误')
            error_box.setText(f'处理失败：{error_msg}')
            error_box.setIcon(QMessageBox.Critical)
            error_box.exec_()
            logging.error(f'处理失败：{error_msg}')

def ensure_directories():
    """确保必要的目录结构存在"""
    required_dirs = ['logs', 'bak', '供应商对账明细']
    for directory in required_dirs:
        if not os.path.exists(directory):
            os.makedirs(directory)
            logging.info(f'创建目录: {directory}')

def check_expiration():
    """
    检查程序是否过期
    
    检查当前日期是否超过2025年12月31日。
    如果超过，则返回False，表示程序已过期；否则返回True。
    
    Returns:
        bool: 程序是否有效（True表示有效，False表示已过期）
    """
    # 获取当前日期
    current_date = datetime.now()
    # 设置过期日期为2025年12月31日
    expiration_date = datetime(2025, 12, 31)
    
    # 如果当前日期超过了2025年12月31日，则程序已过期
    if current_date > expiration_date:
        return False
        
    return True

def main():
    try:
        # 确保必要的目录存在
        ensure_directories()
        
        # 配置日志
        log_filename = os.path.join('logs', f'app_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_filename, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        
        app = QApplication(sys.argv)
        # 导入资源文件并设置全局窗口图标
        import resources
        icon = QIcon(':/icons/app_icon')
        app.setWindowIcon(icon)
        
        # 确保任务栏图标与应用程序图标一致（Windows平台特定）
        if sys.platform == 'win32':
            import ctypes
            app_id = f'MC.ReconUI.{VERSION}'  # 应用程序ID
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(app_id)
        
        # 检查程序是否过期
        if not check_expiration():
            logging.error('程序版本已过期，需要更新')
            error_box = QMessageBox(None)
            error_box.setWindowTitle('版本过期')
            error_box.setText('版本过低，请联系开发者Cayman更新')
            error_box.setIcon(QMessageBox.Critical)
            error_box.exec_()
            sys.exit(1)
        else:
            logging.info('程序版本检查通过')
        
        window = MainWindow()
        window.show()
        logging.info('应用程序启动成功')
        sys.exit(app.exec_())
    except Exception as e:
        logging.error(f'应用程序启动失败: {e}')
        error_box = QMessageBox(None)
        error_box.setWindowTitle('错误')
        error_box.setText(f'应用程序启动失败: {e}')
        error_box.setIcon(QMessageBox.Critical)
        error_box.exec_()
        sys.exit(1)

if __name__ == '__main__':
    main()